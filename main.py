# Projet pour Math en Jeans 2026
# 2 Objectif
#
# n°1 : Lire coordonnées prédifinie dans un tableau excele et l'afficher dans une fenêtre graphique 
# puis placer le centre à l'aide des méthodes
# n°2 : Placement aléatoire de points selon les paramètres donnéss
#
#
# utilisation de libary : 
# simplegraphics-python, openpyxl, scipy.spatial, numpy, math, tkinter
#

import openpyxl
import tkinter
from tkinter import *
from tkinter import messagebox

import moyenneDesPoints
# Importations des modules locaux
from classes import *
from Baricentre import *
from moyenneDesPoints import *
from PointDejaPresent import *
from EnveloppeConvex import *
from ProjectionMethods import *

# --- Configuration ---
screen_width = 1200
screen_height = 900
tableau_excel = "EntreeDesPoints.xlsx"
tableau_list_points = "ListPoint.xlsx"

# Initialisation Excel & Graphique
try:
    File_tableau = openpyxl.load_workbook(tableau_excel)
    File_points = openpyxl.load_workbook(tableau_list_points)
    Sheet_tableau = File_tableau.active
    Sheet_points = File_points.active
except Exception as e:
    print(f"Erreur chargement Excel: {e}")

SimpleGraphics.resize(screen_width, screen_height)
nuage = PointCloud()

# --- Fonctions de l'Interface ---

def AjouterPoints(n_points_str):
    try:
        nombre = int(n_points_str)
        nuage.createRandomPoints(nombre, screen_width, screen_height)
        setup()
    except ValueError:
        messagebox.showerror("Erreur", "Entrez un nombre valide pour les points.")

def AjouterPointsTableau():
    nuage.GetPoints(Sheet_tableau)
    setup()

def ClearScreen():
    SimpleGraphics.clear()
    nuage.List_points.clear()
    Sheet_points.delete_rows(1, Sheet_points.max_row)
    File_points.save(tableau_list_points)

def QuitterProgramme():
    print("Fermeture du programme")
    win.destroy()
    SimpleGraphics.close()

def setup():
    nuage.drawAllPoints()
    nuage.getPositions()

# --- STRUCTURE DYNAMIQUE ---
# Format: ("Nom du bouton", Référence_Fonction, Type_Argument)
# Type_Argument peut être: "INPUT" (lit la box), "NUAGE" (passe l'objet), ou None
functions_list = [
    ("Générer Points", AjouterPoints, "INPUT"),
    ("Charger Excel", AjouterPointsTableau, None),
    ("Calculer Barycentre", baricentre, "NUAGE"),
    ("calculer Moyenne", Moyenne, "NUAGE"),
    ("méthode du point deja présent", PointDejaPresent, "NUAGE"),
    ("Afficher Enveloppe Convexe", EnveloppeConvex, "NUAGE"),
    ("Projection Méthode 1", ProjectionMethod1, "NUAGE"),
    ("Projection Méthode 2", ProjectionMethod2, "NUAGE"),
    ("Effacer Tout", ClearScreen, None),
    ("Quitter", QuitterProgramme, None)
]

class Window(tkinter.Tk):
    def __init__(self, funcs):
        super().__init__()
        self.title("Math en Jeans 2026")
        self.geometry("400x600")

        # UI Header
        Label(self, text="Interface de Contrôle", font=("Arial", 14, "bold")).pack(pady=10)

        # Input Box Global
        input_frame = Frame(self)
        input_frame.pack(pady=10)
        Label(input_frame, text="Valeur (n):").pack(side=LEFT)
        self.input_box = Entry(input_frame, width=10)
        self.input_box.pack(side=LEFT, padx=5)
        self.input_box.insert(0, "10")

        # Boucle de création de boutons
        for name, func, arg_type in funcs:
            if arg_type == "INPUT":
                # On capture la valeur de la box au moment du clic
                cmd = lambda f=func: f(self.input_box.get())
            elif arg_type == "NUAGE":
                # On passe directement l'objet nuage global
                cmd = lambda f=func: f(nuage)
            else:
                # Appel simple sans argument
                cmd = func

            Button(self, text=name, command=cmd, height=2, width=25).pack(pady=5)

if __name__ == "__main__":
    win = Window(functions_list)
    win.mainloop()